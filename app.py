import os
from collections import defaultdict
import json # Import json module
from flask import Flask, render_template, request, redirect, url_for, flash, session, abort, jsonify, send_file, send_from_directory
from functools import wraps
from datetime import datetime, timezone, date
from threading import Thread
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from models import db, User, Lehrer, Angebot, Klasse, Sammelangebot, StundenplanEintrag, Einstellung, Event, KanbanList, KanbanCard, Vertretungsplan, VertretungsplanEintrag, GespeicherterStundenplan, GespeicherterStundenplanEintrag
from flask_mail import Mail, Message # type: ignore
from constants import TAGE_DER_WOCHE, LEHRER_FARBEN_MAP, DEFAULT_LEHRER_FARBE_NAME
from forms import LoginForm
import scheduler_web
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from werkzeug.utils import secure_filename
from itsdangerous import URLSafeTimedSerializer, SignatureExpired
from dotenv import load_dotenv


app = Flask(__name__)

# Lade Umgebungsvariablen aus der .env Datei
load_dotenv()

# --- Konfiguration ---
# HINWEIS: SECRET_KEY sollte in einer .env-Datei oder Umgebungsvariable gespeichert werden
# Wenn du eine .env-Datei verwendest, stelle sicher, dass python-dotenv installiert ist und die .env-Datei geladen wird.
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or 'your_super_secret_key_here' # Wichtig für WTForms und Sessions
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///site.db' # Oder 'sqlite:///database.db', falls du diese Datei verwendest
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# --- Mail Konfiguration (aus Umgebungsvariablen laden) ---
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT'))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'false').lower() in ['true', 'on', '1']
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER')

# Initialisiere Erweiterungen
db.init_app(app) # Initialisiere SQLAlchemy mit der App
bcrypt = Bcrypt(app) # Initialisiere Bcrypt mit der App
login_manager = LoginManager(app) # Initialisiere LoginManager mit der App
mail = Mail(app) # Initialisiere Flask-Mail
login_manager.login_view = 'login' # Die Route, zu der umgeleitet wird, wenn Login erforderlich ist
login_manager.login_message_category = 'info' # Kategorie für Flash-Nachrichten

# --- Token Serializer für Passwort-Reset ---
s = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# --- Rollen-Definitionen ---
ROLE_ADMIN = 'Admin'
ROLE_PLANER = 'Planer'
ROLE_BENUTZER = 'Benutzer'

# --- Rollen-basierter Zugriffs-Decorator ---
def role_required(allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized() # Redirect to login page
            if current_user.role not in allowed_roles:
                flash('Sie haben keine Berechtigung, auf diese Seite zuzugreifen.', 'danger')
                abort(403) # Forbidden
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- Hilfsfunktion für asynchronen E-Mail-Versand ---
def send_async_email(app, msg):
    with app.app_context():
        try:
            mail.send(msg)
        except Exception as e:
            print(f"Fehler beim Senden der E-Mail: {e}")

def send_email(subject, recipients, html_body, text_body):
    """Erstellt und versendet eine E-Mail in einem separaten Thread."""
    if not app.config.get('MAIL_USERNAME') or not app.config.get('MAIL_PASSWORD'):
        print("WARNUNG: MAIL_USERNAME oder MAIL_PASSWORD nicht konfiguriert. E-Mail wird nicht gesendet.")
        return
    msg = Message(subject, recipients=recipients, body=text_body, html=html_body)
    Thread(target=send_async_email, args=(app, msg)).start()

# --- Routen ---
@app.route("/")
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER, ROLE_BENUTZER])
def stundenplan_anzeigen():
    # --- GEÄNDERTE LOGIK: Es wird JEDER existierende Vertretungsplan geholt ---
    # Da es immer nur einen geben kann, ist .first() korrekt.
    vorhandener_vertretungsplan = Vertretungsplan.query.first()

    # Das Flag is_vertretungsplan ist wahr, wenn irgendein Plan existiert.
    is_vertretungsplan = bool(vorhandener_vertretungsplan)
    # --- ÄNDERUNG: Dashboard zeigt IMMER den normalen Plan an ---
    woche = request.args.get('woche', 'A')
    title_woche_info = f" (Woche {woche})"

    plan_eintraege = StundenplanEintrag.query.filter_by(woche=woche).all()

    klassen_reihenfolge_a_str = get_setting('klassen_reihenfolge_a', '')
    klassen_reihenfolge_b_str = get_setting('klassen_reihenfolge_b', '')
    alle_klassen_map = {k.name: k for k in Klasse.query.all()}
    
    # --- ÄNDERUNG: Klassenfilterung ist immer für A/B Woche, nicht mehr für Vertretungsplan ---
    if woche == 'A':
        geordnete_klassen = [alle_klassen_map[name] for name in klassen_reihenfolge_a_str.split(',') if name in alle_klassen_map]
    else:
        geordnete_klassen = [alle_klassen_map[name] for name in klassen_reihenfolge_b_str.split(',') if name in alle_klassen_map]

    zeiten_text = get_setting('zeiten_text', '08:00-08:45\n08:45-09:30\n09:30-09:45 Pause\n09:45-10:30\n10:30-11:15\n11:15-11:45 Pause\n11:45-12:30\n12:30-13:15')
    zeit_slots = parse_zeiten(zeiten_text)
    num_slots = len(zeit_slots)

    plan_data_per_tag = defaultdict(lambda: defaultdict(dict))
    for eintrag in plan_eintraege:
        if not all([eintrag.angebot, eintrag.lehrer1, eintrag.klasse]):
            continue

        eintrag_dict = {
            'angebot': {'id': eintrag.angebot.id, 'name': eintrag.angebot.name},
            'lehrer1': {'id': eintrag.lehrer1.id, 'name': eintrag.lehrer1.name, 'farbe': eintrag.lehrer1.farbe},
            'lehrer2': {'id': eintrag.lehrer2.id, 'name': eintrag.lehrer2.name, 'farbe': eintrag.lehrer2.farbe} if eintrag.lehrer2 else None
        }
        if eintrag.slot < num_slots:
            plan_data_per_tag[eintrag.tag][eintrag.slot][eintrag.klasse.name] = eintrag_dict

    title = f"Stundenplan{title_woche_info}"

    alle_lehrer = db.session.query(Lehrer).all()

    return render_template(
        'stundenplan.html',
        title=title,
        geordnete_klassen=geordnete_klassen,
        plan_data_per_tag=plan_data_per_tag,
        zeit_slots=zeit_slots,
        tage_der_woche=TAGE_DER_WOCHE,
        woche=woche,
        alle_lehrer=alle_lehrer,
        is_vertretungsplan=is_vertretungsplan, # Flag wird weiterhin übergeben
        aktiver_vertretungsplan=vorhandener_vertretungsplan # Plan-Objekt wird übergeben
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('stundenplan_anzeigen'))
    form = LoginForm() # Use the updated LoginForm
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first() # Authenticate by username
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            login_user(user, remember=form.remember.data)
            next_page = request.args.get('next')
            flash('Anmeldung erfolgreich!', 'success')
            return redirect(next_page) if next_page else redirect(url_for('stundenplan_anzeigen'))
        else:
            flash('Anmeldung fehlgeschlagen. Bitte Benutzername und Passwort prüfen.', 'danger')
    return render_template('login.html', title='Anmelden', form=form)

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password_request():
    if current_user.is_authenticated:
        return redirect(url_for('stundenplan_anzeigen'))
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()
        if user:
            token = s.dumps(user.email, salt='password-reset-salt')
            reset_url = url_for('reset_token', token=token, _external=True)
            html = render_template('email/reset_password.html', reset_url=reset_url)
            text = render_template('email/reset_password.txt', reset_url=reset_url)
            send_email("Passwort zurücksetzen für Stundentool", [user.email], html, text)
        flash('Wenn ein Konto mit dieser E-Mail existiert, wurde eine Anleitung zum Zurücksetzen des Passworts gesendet.', 'info')
        return redirect(url_for('login'))
    return render_template('request_reset.html', title='Passwort zurücksetzen')

@app.route("/reset_password/<token>", methods=['GET', 'POST'])
def reset_token(token):
    if current_user.is_authenticated:
        return redirect(url_for('stundenplan_anzeigen'))
    try:
        email = s.loads(token, salt='password-reset-salt', max_age=1800) # 30 Minuten gültig
    except SignatureExpired:
        flash('Der Link zum Zurücksetzen des Passworts ist abgelaufen.', 'danger')
        return redirect(url_for('reset_password_request'))
    except Exception:
        flash('Der Link zum Zurücksetzen des Passworts ist ungültig.', 'danger')
        return redirect(url_for('reset_password_request'))
    
    user = User.query.filter_by(email=email).first_or_404()

    if request.method == 'POST':
        password = request.form.get('password')
        if password:
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
            user.password = hashed_password
            db.session.commit()
            flash('Dein Passwort wurde erfolgreich aktualisiert. Du kannst dich jetzt anmelden.', 'success')
            return redirect(url_for('login'))
        else:
            flash('Das Passwort darf nicht leer sein.', 'danger')

    return render_template('reset_token.html', title='Neues Passwort festlegen', token=token)

@app.route('/logout')
def logout():
    logout_user()
    flash('Sie wurden abgemeldet.', 'info')
    return redirect(url_for('login')) # Weiterleitung zur Login-Seite nach dem Abmelden

# =====================================================================
# Lehrer-Verwaltung
# =====================================================================

@app.route('/lehrer')
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def lehrer_verwalten():
    lehrer_liste = Lehrer.query.order_by(Lehrer.name).all()
    alle_angebote = Angebot.query.order_by(Angebot.name).all()
    klassen_liste = Klasse.query.order_by(Klasse.name).all()
    farben_reverse_map = {v: k for k, v in LEHRER_FARBEN_MAP.items()}
    lehrer_hauptangebot = {}
    lehrer_hauptangebot_id = {}
    lehrer_einsatz_klassen = {}
    klassen_id_name_map = {k.id: k.name for k in klassen_liste}
    for l in lehrer_liste:
        hauptangebot_name = ""
        hauptangebot_id = None
        if hasattr(l, "hauptangebot") and l.hauptangebot:
            hauptangebot_name = l.hauptangebot.name
            hauptangebot_id = l.hauptangebot.id
        elif hasattr(l, "hauptangebot_id") and l.hauptangebot_id:
            angebot = Angebot.query.get(l.hauptangebot_id)
            if angebot:
                hauptangebot_name = angebot.name
                hauptangebot_id = angebot.id
        lehrer_hauptangebot[l.id] = hauptangebot_name
        lehrer_hauptangebot_id[l.id] = hauptangebot_id
        # NEU: Einsatzklassen als Liste von Namen für Anzeige
        einsatz_klassen_ids = []
        raw_einsatz = getattr(l, "einsatz_klassen", [])
        # --- PATCH: Immer als Liste speichern und laden, auch für SQLite/Text ---
        # Versuche, das Feld als Liste zu interpretieren (PickleType, JSON, Text)
        if isinstance(raw_einsatz, str):
            try:
                # Falls als JSON-String gespeichert
                einsatz_klassen_ids = json.loads(raw_einsatz)
                if not isinstance(einsatz_klassen_ids, list):
                    einsatz_klassen_ids = []
            except Exception:
                einsatz_klassen_ids = []
        elif isinstance(raw_einsatz, list):
            einsatz_klassen_ids = raw_einsatz
        elif raw_einsatz is None:
            einsatz_klassen_ids = []
        else:
            try:
                einsatz_klassen_ids = list(raw_einsatz)
            except Exception:
                einsatz_klassen_ids = []
        # --- PATCH ENDE ---
        einsatz_klassen_namen = [klassen_id_name_map.get(cid, str(cid)) for cid in einsatz_klassen_ids] if einsatz_klassen_ids else []
        lehrer_einsatz_klassen[l.id] = einsatz_klassen_namen
    return render_template('lehrer_verwalten.html',
                           title='Lehrer verwalten',
                           lehrer_liste=lehrer_liste,
                           alle_angebote=alle_angebote,
                           klassen_liste=klassen_liste,
                           farben_reverse_map=farben_reverse_map,
                           tage_der_woche=TAGE_DER_WOCHE,
                           LEHRER_FARBEN_MAP=LEHRER_FARBEN_MAP,
                           DEFAULT_LEHRER_FARBE_NAME=DEFAULT_LEHRER_FARBE_NAME,
                           lehrer_hauptangebot=lehrer_hauptangebot,
                           lehrer_hauptangebot_id=lehrer_hauptangebot_id,
                           lehrer_einsatz_klassen=lehrer_einsatz_klassen,
                           klassen_id_name_map=klassen_id_name_map)

@app.route('/lehrer/save', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def lehrer_save():
    lehrer_id = request.form.get('lehrer_id')
    name = request.form.get('name')
    gesamt_stunden_soll_input = request.form.get('stunden_gesamt_soll_input')
    stunden_a_input = request.form.get('stunden_a_input')
    stunden_b_input = request.form.get('stunden_b_input')
    gewaehlte_farbe_name = request.form.get('farbe')
    tage = request.form.getlist('tage')
    max_stunden_pro_tag = int(request.form.get('max_stunden_pro_tag', 8))
    gewaehlte_angebote_ids = [int(x) for x in request.form.getlist('angebote') if x.isdigit()]
    hauptangebot_id = request.form.get('hauptangebot')
    einsatz_klassen_ids = [int(x) for x in request.form.getlist('einsatz_klassen') if x.isdigit()]  # NEU

    # --- NEU: Wenn Gesamtstunden gesetzt, für beide Wochen übernehmen ---
    if gesamt_stunden_soll_input:
        stunden_a_input = gesamt_stunden_soll_input
        stunden_b_input = gesamt_stunden_soll_input

    # --- NEU: Wenn nur eine Woche ausgefüllt, andere übernehmen ---
    if stunden_a_input and not stunden_b_input:
        stunden_b_input = stunden_a_input
    if stunden_b_input and not stunden_a_input:
        stunden_a_input = stunden_b_input

    # Konvertiere Stunden-Inputs zu Integer
    try:
        effektive_stunden_a = int(stunden_a_input) if stunden_a_input else 0
    except ValueError:
        effektive_stunden_a = 0
    try:
        effektive_stunden_b = int(stunden_b_input) if stunden_b_input else 0
    except ValueError:
        effektive_stunden_b = 0

    # Konvertiere max_stunden_pro_tag von einzelnem Integer zu JSON-Struktur
    max_stunden_json = {}
    for tag in TAGE_DER_WOCHE:
        # Setzt den Wert für alle Tage und Wochen A/B auf den eingegebenen Wert
        max_stunden_json[tag] = {"A": max_stunden_pro_tag, "B": max_stunden_pro_tag}

    if not name: # Diese Zeile war zuvor falsch eingerückt
        flash('Der Name des Lehrers darf nicht leer sein.', 'danger')
        return redirect(url_for('lehrer_verwalten'))
    
    existing_lehrer = Lehrer.query.filter(Lehrer.name == name, Lehrer.id != lehrer_id).first()
    if existing_lehrer:
        flash(f'Ein anderer Lehrer mit dem Namen "{name}" existiert bereits.', 'warning')
        return redirect(url_for('lehrer_verwalten'))

    if lehrer_id:
        lehrer = Lehrer.query.get_or_404(lehrer_id)
        lehrer.name = name
        lehrer.stunden_gesamt_soll_input = gesamt_stunden_soll_input
        lehrer.stunden_a_input = stunden_a_input
        lehrer.stunden_b_input = stunden_b_input
        lehrer.stunden_a = effektive_stunden_a
        lehrer.stunden_b = effektive_stunden_b
        lehrer.farbe = LEHRER_FARBEN_MAP.get(gewaehlte_farbe_name, LEHRER_FARBEN_MAP[DEFAULT_LEHRER_FARBE_NAME]) # type: ignore
        lehrer.tage = tage
        lehrer.max_stunden_pro_tag = max_stunden_json # Speichere als JSON
        # --- Fix: Hauptangebot speichern, auch wenn None (setzt auf NULL in DB) ---
        if hauptangebot_id:
            lehrer.hauptangebot_id = int(hauptangebot_id)
        else:
            lehrer.hauptangebot_id = None

        # Angebote aktualisieren (Many-to-Many)
        lehrer.angebote.clear() # Alle bestehenden entfernen
        for angebot_id in gewaehlte_angebote_ids:
            angebot = Angebot.query.get(angebot_id)
            if angebot:
                lehrer.angebote.append(angebot) # Neue hinzufügen
        
        # NEU: Einsatzklassen speichern (als Liste von IDs, None/[] = alle)
        # --- PATCH: Immer als JSON-String speichern, falls Feld kein echtes JSON/PickleType ist ---
        # Prüfe, ob das Feld in der DB als Text/JSON/PickleType ist
        import sqlalchemy
        col_type = None
        try:
            col_type = type(Lehrer.__table__.columns['einsatz_klassen'].type)
        except Exception:
            pass
        # Wenn PickleType oder JSON, speichere als Liste, sonst als JSON-String
        if col_type and (col_type.__name__ in ['PickleType', 'JSON']):
            lehrer.einsatz_klassen = einsatz_klassen_ids if einsatz_klassen_ids else []
        else:
            lehrer.einsatz_klassen = json.dumps(einsatz_klassen_ids if einsatz_klassen_ids else [])

        flash(f'Lehrer "{name}" wurde erfolgreich aktualisiert.', 'success')
    else:
        lehrer = Lehrer(
            name=name,
            stunden_gesamt_soll_input=gesamt_stunden_soll_input,
            stunden_a_input=stunden_a_input,
            stunden_b_input=stunden_b_input,
            stunden_a=effektive_stunden_a,
            stunden_b=effektive_stunden_b,
            farbe=LEHRER_FARBEN_MAP.get(gewaehlte_farbe_name, LEHRER_FARBEN_MAP[DEFAULT_LEHRER_FARBE_NAME]), # type: ignore
            tage=tage,
            max_stunden_pro_tag=max_stunden_json, # Speichere als JSON
            abwesend_an_tagen=[], # Neue Lehrer starten ohne Abwesenheiten
            # hauptangebot_id NICHT im Konstruktor, sondern nachträglich setzen!
        )
        # Hauptangebot nachträglich setzen, falls vorhanden
        if hauptangebot_id:
            lehrer.hauptangebot_id = int(hauptangebot_id)
        else:
            lehrer.hauptangebot_id = None
        # Angebote hinzufügen
        for angebot_id in gewaehlte_angebote_ids:
            angebot = Angebot.query.get(angebot_id)
            if angebot:
                lehrer.angebote.append(angebot)

        import sqlalchemy
        col_type = None
        try:
            col_type = type(Lehrer.__table__.columns['einsatz_klassen'].type)
        except Exception:
            pass
        if col_type and (col_type.__name__ in ['PickleType', 'JSON']):
            lehrer.einsatz_klassen = einsatz_klassen_ids if einsatz_klassen_ids else []
        else:
            lehrer.einsatz_klassen = json.dumps(einsatz_klassen_ids if einsatz_klassen_ids else [])  # NEU

        db.session.add(lehrer)
    db.session.commit()

    if not lehrer_id:
        flash(f'Lehrer "{name}" wurde erfolgreich hinzugefügt.', 'success')
    return redirect(url_for("lehrer_verwalten"))

@app.route('/lehrer/delete/<int:lehrer_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def lehrer_delete(lehrer_id):
    lehrer = Lehrer.query.get_or_404(lehrer_id)
    db.session.delete(lehrer)
    db.session.commit()
    flash(f'Lehrer "{lehrer.name}" wurde erfolgreich gelöscht.', 'success')
    return redirect(url_for('lehrer_verwalten'))

# =====================================================================
# Klassen-Verwaltung
# =====================================================================

@app.route('/klassen')
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def klassen_verwalten():
    klassen_liste = Klasse.query.order_by(Klasse.name).all()
    alle_angebote = Angebot.query.order_by(Angebot.name).all()
    return render_template('klassen_verwalten.html',
                           title='Klassen verwalten',
                           klassen_liste=klassen_liste,
                           alle_angebote=alle_angebote,
                           tage_der_woche=TAGE_DER_WOCHE)

@app.route('/klassen/save', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def klasse_save():
    klasse_id = request.form.get('klasse_id')
    name = request.form.get('name')
    max_stunden_str = request.form.get('max_stunden_klasse')
    woche = request.form.get('woche')
    arbeitstage = request.form.getlist('arbeitstage')

    # Sammle die Angebotsstunden aus dem Formular
    angebote_stunden_data = []
    alle_angebote = Angebot.query.all()
    for angebot in alle_angebote:
        stunden_gesamt_str = request.form.get(f'angebot_{angebot.id}_gesamt')
        stunden_teilung_str = request.form.get(f'angebot_{angebot.id}_teilung')

        stunden_gesamt = int(stunden_gesamt_str) if stunden_gesamt_str and stunden_gesamt_str.isdigit() else 0

        if stunden_gesamt > 0:
            stunden_teilung = int(stunden_teilung_str) if stunden_teilung_str and stunden_teilung_str.isdigit() else 0
            eintrag = {
                "angebot": angebot.name,
                "stunden_gesamt": stunden_gesamt,
                "stunden_teilung": stunden_teilung
            }
            angebote_stunden_data.append(eintrag)

    if not name:
        flash('Der Klassenname darf nicht leer sein.', 'danger')
        return redirect(url_for('klassen_verwalten'))
    
    existing_klasse_query = Klasse.query.filter(Klasse.name == name)
    if klasse_id:
        existing_klasse_query = existing_klasse_query.filter(Klasse.id != int(klasse_id))
    existing_klasse = existing_klasse_query.first()

    if existing_klasse:
        flash(f'Eine andere Klasse mit dem Namen "{name}" existiert bereits.', 'warning')
        return redirect(url_for('klassen_verwalten'))

    try:
        max_stunden = int(max_stunden_str) if max_stunden_str else 6
    except ValueError:
        flash('Max. Stunden pro Tag muss eine Zahl sein.', 'danger')
        return redirect(url_for('klassen_verwalten'))

    if klasse_id:
        klasse = Klasse.query.get_or_404(klasse_id)
        klasse.name = name
        klasse.max_stunden_klasse = max_stunden
        klasse.woche = woche
        klasse.arbeitstage = arbeitstage
        klasse.angebote_stunden = angebote_stunden_data
        flash(f'Klasse "{name}" wurde aktualisiert.', 'success')
    else:
        neue_klasse = Klasse(
            name=name,
            max_stunden_klasse=max_stunden,
            woche=woche,
            arbeitstage=arbeitstage,
            angebote_stunden=angebote_stunden_data
        )
        db.session.add(neue_klasse)
        flash(f'Klasse "{name}" wurde hinzugefügt.', 'success')
    
    db.session.commit()
    return redirect(url_for('klassen_verwalten'))

@app.route('/klassen/delete/<int:klasse_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def klasse_delete(klasse_id):
    klasse = Klasse.query.get_or_404(klasse_id)
    db.session.delete(klasse)
    db.session.commit()
    flash(f'Klasse "{klasse.name}" wurde gelöscht.', 'success')
    return redirect(url_for('klassen_verwalten'))

# =====================================================================
# Abwesenheiten
# =====================================================================

@app.route('/abwesenheiten', methods=['GET', 'POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def abwesenheiten_verwalten():
    if request.method == 'POST':
        lehrer_id = request.form.get('lehrer_id')
        if not lehrer_id:
            flash('Kein Lehrer ausgewählt.', 'warning')
            return redirect(url_for('abwesenheiten_verwalten'))

        lehrer = Lehrer.query.get_or_404(lehrer_id)
        abwesend_an = request.form.getlist('abwesend_an')
        lehrer.abwesend_an_tagen = abwesend_an
        db.session.commit()
        flash(f'Abwesenheiten für {lehrer.name} gespeichert.', 'success')
        # Redirect back to the same teacher's view
        return redirect(url_for('abwesenheiten_verwalten', lehrer_id=lehrer_id))

    # GET request handling
    lehrer_liste = Lehrer.query.order_by(Lehrer.name).all()
    selected_lehrer_id = request.args.get('lehrer_id', type=int)
    selected_lehrer = None
    if selected_lehrer_id:
        selected_lehrer = Lehrer.query.get(selected_lehrer_id)

    return render_template('abwesenheiten_verwalten.html', title='Abwesenheiten', lehrer_liste=lehrer_liste, tage_der_woche=TAGE_DER_WOCHE, selected_lehrer=selected_lehrer)

# =====================================================================
# Übersichten & Planerstellung
# =====================================================================

@app.route('/verwaltung/stundenplan', endpoint='stundenplan_verwaltung')
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def stundenplan_verwalten():
    gespeicherte_plaene = GespeicherterStundenplan.query.order_by(GespeicherterStundenplan.created_at.desc()).all()
    woche = request.args.get('woche', 'A')
    klassen_reihenfolge_a_str = get_setting('klassen_reihenfolge_a', '')
    klassen_reihenfolge_b_str = get_setting('klassen_reihenfolge_b', '')
    alle_klassen_map = {k.name: k for k in Klasse.query.all()}
    if woche == 'A':
        geordnete_klassen = [alle_klassen_map[name] for name in klassen_reihenfolge_a_str.split(',') if name in alle_klassen_map]
    else:
        geordnete_klassen = [alle_klassen_map[name] for name in klassen_reihenfolge_b_str.split(',') if name in alle_klassen_map]
    zeiten_text = get_setting('zeiten_text', '08:00-08:45\n08:45-09:30\n09:30-09:45 Pause\n09:45-10:30\n10:30-11:15\n11:15-11:45 Pause\n11:45-12:30\n12:30-13:15')
    zeit_slots = parse_zeiten(zeiten_text)
    num_slots = len(zeit_slots)
    plan_eintraege = StundenplanEintrag.query.filter_by(woche=woche).all()
    plan_data_per_tag = defaultdict(lambda: defaultdict(dict))
    for eintrag in plan_eintraege:
        if not all([eintrag.angebot, eintrag.lehrer1, eintrag.klasse]):
            continue
        eintrag_dict = {
            'angebot': {'id': eintrag.angebot.id, 'name': eintrag.angebot.name},
            'lehrer1': {'id': eintrag.lehrer1.id, 'name': eintrag.lehrer1.name, 'farbe': eintrag.lehrer1.farbe},
            'lehrer2': {'id': eintrag.lehrer2.id, 'name': eintrag.lehrer2.name, 'farbe': eintrag.lehrer2.farbe} if eintrag.lehrer2 else None
        }
        if eintrag.slot < num_slots:
            plan_data_per_tag[eintrag.tag][eintrag.slot][eintrag.klasse.name] = eintrag_dict
    scheduler_meldungen = session.pop('scheduler_meldungen', None)
    return render_template(
        'stundenplan_verwaltung.html',
        title='Stundenplan Verwaltung',
        gespeicherte_plaene=gespeicherte_plaene,
        scheduler_meldungen=scheduler_meldungen,
        geordnete_klassen=geordnete_klassen,
        plan_data_per_tag=plan_data_per_tag,
        zeit_slots=zeit_slots,
        tage_der_woche=TAGE_DER_WOCHE,
        woche=woche,
        alle_lehrer=Lehrer.query.order_by(Lehrer.name).all(),
        alle_angebote=Angebot.query.order_by(Angebot.name).all(),
        klassen_liste=Klasse.query.order_by(Klasse.name).all(),  # falls im Template benötigt
    )

@app.route('/stundenplan/speichern', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def stundenplan_speichern():
    plan_name = request.form.get('plan_name')
    if not plan_name:
        flash('Ein Name für den Plan ist erforderlich.', 'danger')
        return redirect(url_for('stundenplan_verwaltung'))

    if GespeicherterStundenplan.query.filter_by(name=plan_name).first():
        flash(f'Ein Plan mit dem Namen "{plan_name}" existiert bereits. Bitte wählen Sie einen anderen Namen.', 'warning')
        return redirect(url_for('stundenplan_verwaltung'))

    try:
        # 1. Neuen Container für den gespeicherten Plan erstellen
        neuer_gespeicherter_plan = GespeicherterStundenplan(name=plan_name)
        db.session.add(neuer_gespeicherter_plan)
        db.session.flush() # Nötig, um die ID für die Einträge zu bekommen

        # 2. Alle Einträge aus dem aktiven Plan kopieren
        aktive_eintraege = StundenplanEintrag.query.all()
        if not aktive_eintraege:
            flash('Der aktuelle Stundenplan ist leer. Nichts zu speichern.', 'info')
            return redirect(url_for('stundenplan_verwaltung'))

        for eintrag in aktive_eintraege:
            kopie = GespeicherterStundenplanEintrag(
                gespeicherter_stundenplan_id=neuer_gespeicherter_plan.id,
                woche=eintrag.woche,
                tag=eintrag.tag,
                slot=eintrag.slot,
                klasse_id=eintrag.klasse_id,
                angebot_id=eintrag.angebot_id,
                lehrer1_id=eintrag.lehrer1_id,
                lehrer2_id=eintrag.lehrer2_id
            )
            db.session.add(kopie)
        
        db.session.commit()
        flash(f'Stundenplan als "{plan_name}" erfolgreich gespeichert.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Speichern des Stundenplans: {e}', 'danger')

    return redirect(url_for('stundenplan_verwaltung'))

@app.route('/stundenplan/laden', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def stundenplan_laden():
    plan_id = request.form.get('plan_id')
    if not plan_id:
        flash('Kein Plan zum Laden ausgewählt.', 'danger')
        return redirect(url_for('stundenplan_verwaltung'))

    plan_zum_laden = GespeicherterStundenplan.query.get_or_404(plan_id)

    try:
        # 1. Aktiven Plan löschen
        StundenplanEintrag.query.delete()

        # 2. Gespeicherten Plan in den aktiven Plan kopieren
        for eintrag in plan_zum_laden.eintraege:
            neuer_aktiver_eintrag = StundenplanEintrag(
                woche=eintrag.woche,
                tag=eintrag.tag,
                slot=eintrag.slot,
                klasse_id=eintrag.klasse_id,
                angebot_id=eintrag.angebot_id,
                lehrer1_id=eintrag.lehrer1_id,
                lehrer2_id=eintrag.lehrer2_id
            )
            db.session.add(neuer_aktiver_eintrag)
        
        db.session.commit()
        flash(f'Stundenplan "{plan_zum_laden.name}" wurde erfolgreich geladen.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Laden des Stundenplans: {e}', 'danger')

    return redirect(url_for('stundenplan_verwaltung'))

@app.route('/stundenplan/gespeichert/loeschen/<int:plan_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def stundenplan_gespeichert_loeschen(plan_id):
    plan = GespeicherterStundenplan.query.get_or_404(plan_id)
    try:
        db.session.delete(plan)
        db.session.commit()
        flash(f'Gespeicherter Plan "{plan.name}" wurde gelöscht.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Löschen des gespeicherten Plans: {e}', 'danger')
    
    return redirect(url_for('stundenplan_verwaltung'))

@app.route('/stundenplan/leeren', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def stundenplan_leeren():
    """Löscht alle Einträge aus dem aktiven, bearbeitbaren Stundenplan."""
    try:
        num_deleted = db.session.query(StundenplanEintrag).delete()
        db.session.commit()
        flash(f'Der Stundenplan wurde geleert. {num_deleted} Einträge wurden gelöscht.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Leeren des Stundenplans: {e}', 'danger')
    
    return redirect(url_for('stundenplan_verwaltung'))

@app.route('/plan/erstellen', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def plan_erstellen():
    # NEU: Modus aus dem Formular lesen
    modus = request.form.get('modus', 'neu')

    # NEU: Bestehende Einträge sammeln, wenn Modus 'behalten' ist
    existing_entries = []
    if modus == 'behalten':
        flash('Modus "Freie Slots füllen" gewählt. Bestehende Einträge werden beibehalten.', 'info')
        # Lade den bestehenden Plan und formatiere ihn für den Scheduler
        eintraege_db = StundenplanEintrag.query.options(
            db.joinedload(StundenplanEintrag.klasse),
            db.joinedload(StundenplanEintrag.angebot),
            db.joinedload(StundenplanEintrag.lehrer1),
            db.joinedload(StundenplanEintrag.lehrer2)
        ).all()
        for e in eintraege_db:
            if not all([e.klasse, e.angebot, e.lehrer1]): continue
            existing_entries.append({
                'woche': e.woche,
                'tag': e.tag,
                'slot': e.slot,
                'klasse_name': e.klasse.name,
                'angebot_name': e.angebot.name,
                'lehrer_name': e.lehrer1.name,
                'lehrer2_name': e.lehrer2.name if e.lehrer2 else None
            })
    else: # modus == 'neu'
        flash('Modus "Alles neu planen" gewählt. Der bestehende Plan wird gelöscht.', 'info')

    flash('Starte Planerstellung... Dies kann einen Moment dauern.', 'info')

    # --- Sammelangebote robust laden ---
    sammelangebote_db = Sammelangebot.query.all()
    sammelangebote_liste_dict = []
    for sa in sammelangebote_db:
        kernangebot_name = sa.kernangebot.name if sa.kernangebot else None
        lehrer_namen = []
        if hasattr(sa, "lehrer") and sa.lehrer:
            lehrer_namen = [sa.lehrer.name]
        teilnehmende_klassen_namen = [k.name for k in sa.teilnehmende_klassen] if hasattr(sa, "teilnehmende_klassen") else []
        sammelangebote_liste_dict.append({
            "id": sa.id,
            "name": sa.name,
            "kernangebot": kernangebot_name,
            "lehrer_namen": lehrer_namen,
            "dauer_stunden": getattr(sa, "dauer_stunden", 2),
            "woche_typ": getattr(sa, "woche_typ", "AB"),
            "teilnehmende_klassen_namen": teilnehmende_klassen_namen
        })

    # --- Hole alle Daten VOR der Prüfung ---
    lehrer_db = Lehrer.query.all()
    lehrerliste_dict = [
        {
            "name": l.name,
            "stunden_a": l.stunden_a,
            "stunden_b": l.stunden_b,
            "tage": l.tage,
            "abwesend_an_tagen": l.abwesend_an_tagen,
            "max_stunden_pro_tag": l.max_stunden_pro_tag,
            "angebote": [{"angebot": a.name, "hauptangebot": (l.hauptangebot_id == a.id if hasattr(l, 'hauptangebot_id') else False)} for a in l.angebote],
            "einsatz_klassen": getattr(l, "einsatz_klassen", [])  # NEU: Liste von Klassen-IDs (oder [])
        } for l in lehrer_db
    ]
    klassen_db = Klasse.query.all()
    klassenliste_dict = [
        {
            "klasse": k.name,
            "woche": k.woche,
            "arbeitstage": k.arbeitstage,
            "max_stunden_klasse": k.max_stunden_klasse,
            "angebote_stunden": k.angebote_stunden
        } for k in klassen_db
    ]
    angebote_db = Angebot.query.all()
    angebote_liste_dict = [{"name": a.name, "nur_ein_doppelblock_pro_tag": a.nur_ein_doppelblock_pro_tag, "block_groesse": a.block_groesse} for a in angebote_db]
    klassen_reihenfolge_a = get_setting('klassen_reihenfolge_a', '').split(',')
    klassen_reihenfolge_b = get_setting('klassen_reihenfolge_b', '').split(',')
    klassen_namen_im_plan_display_order = list(dict.fromkeys(klassen_reihenfolge_a + klassen_reihenfolge_b))
    stunden_pro_tag_str = get_setting('stunden_pro_tag', '{"Mo": 6, "Di": 6, "Mi": 6, "Do": 6, "Fr": 6}')
    stunden_pro_tag_config = json.loads(stunden_pro_tag_str)
    num_slots_pro_tag = max(stunden_pro_tag_config.values()) if stunden_pro_tag_config else 8

    # --- Prüfe, ob überhaupt Anforderungen existieren ---
    if not klassenliste_dict or not lehrerliste_dict:
        flash("Es sind keine Klassen oder Lehrer für die Planerstellung vorhanden.", "danger")
        return redirect(url_for('stundenplan_verwaltung'))
    if not any(k["angebote_stunden"] for k in klassenliste_dict):
        flash("Keine Angebote für die Klassen definiert. Bitte ergänzen!", "danger")
        return redirect(url_for('stundenplan_verwaltung'))

    lehrer_db = Lehrer.query.all()
    lehrerliste_dict = [
        {
            "name": l.name,
            "stunden_a": l.stunden_a,
            "stunden_b": l.stunden_b,
            "tage": l.tage,
            "abwesend_an_tagen": l.abwesend_an_tagen,
            "max_stunden_pro_tag": l.max_stunden_pro_tag,
            "angebote": [{"angebot": a.name, "hauptangebot": (l.hauptangebot_id == a.id if hasattr(l, 'hauptangebot_id') else False)} for a in l.angebote],
            "einsatz_klassen": getattr(l, "einsatz_klassen", [])  # NEU: Liste von Klassen-IDs (oder [])
        } for l in lehrer_db
    ]
    klassen_db = Klasse.query.all()
    klassenliste_dict = [
        {
            "klasse": k.name,
            "woche": k.woche,
            "arbeitstage": k.arbeitstage,
            "max_stunden_klasse": k.max_stunden_klasse,
            "angebote_stunden": k.angebote_stunden
        } for k in klassen_db
    ]
    angebote_db = Angebot.query.all()
    angebote_liste_dict = [{"name": a.name, "nur_ein_doppelblock_pro_tag": a.nur_ein_doppelblock_pro_tag, "block_groesse": a.block_groesse} for a in angebote_db]
    klassen_reihenfolge_a = get_setting('klassen_reihenfolge_a', '').split(',')
    klassen_reihenfolge_b = get_setting('klassen_reihenfolge_b', '').split(',')
    klassen_namen_im_plan_display_order = list(dict.fromkeys(klassen_reihenfolge_a + klassen_reihenfolge_b))
    stunden_pro_tag_str = get_setting('stunden_pro_tag', '{"Mo": 6, "Di": 6, "Mi": 6, "Do": 6, "Fr": 6}')
    stunden_pro_tag_config = json.loads(stunden_pro_tag_str)
    num_slots_pro_tag = max(stunden_pro_tag_config.values()) if stunden_pro_tag_config else 8

    # --- Prüfe, ob überhaupt Anforderungen existieren ---
    if not klassenliste_dict or not lehrerliste_dict:
        flash("Es sind keine Klassen oder Lehrer für die Planerstellung vorhanden.", "danger")
        return redirect(url_for('stundenplan_verwaltung'))
    if not any(k["angebote_stunden"] for k in klassenliste_dict):
        flash("Keine Angebote für die Klassen definiert. Bitte ergänzen!", "danger")
        return redirect(url_for('stundenplan_verwaltung'))

    try:
        # Debug-Ausgabe für die wichtigsten Eingabedaten
        print("DEBUG: Starte Planerstellung mit folgenden Parametern:")
        print("Lehrer:", lehrerliste_dict)
        print("Klassen:", klassenliste_dict)
        print("Angebote:", angebote_liste_dict)
        print("Sammelangebote:", sammelangebote_liste_dict)
        print("Slots pro Tag:", num_slots_pro_tag)
        print("Stunden pro Tag Config:", stunden_pro_tag_config)
        plan_rohdaten, meldungen = scheduler_web.generate_schedule_data(
            lehrerliste_dict, klassenliste_dict, angebote_liste_dict, sammelangebote_liste_dict,
            klassen_namen_im_plan_display_order, num_slots_pro_tag, stunden_pro_tag_config,
            existing_entries=existing_entries # NEU
        )
        print("DEBUG: Planerstellung abgeschlossen.")
    except Exception as e:
        import traceback
        print("Fehler bei der Planerstellung:", e)
        traceback.print_exc()
        flash(f'Fehler bei der Planerstellung: {e}', 'danger')
        return redirect(url_for('stundenplan_verwaltung'))

    # --- ALTEN PLAN LÖSCHEN und NEUEN SPEICHERN ---
    # Dies geschieht jetzt immer, da der Scheduler einen kompletten Plan zurückgibt
    try:
        num_deleted = db.session.query(StundenplanEintrag).delete()
        db.session.commit()
        if num_deleted > 0 and modus == 'neu':
             flash(f'{num_deleted} alte Planeinträge gelöscht.', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Löschen des alten Plans: {e}', 'danger')
        return redirect(url_for('stundenplan_verwaltung'))

    klassen_map = {k.name: k for k in klassen_db}
    angebote_map = {a.name: a for a in angebote_db}
    lehrer_map = {l.name: l for l in lehrer_db}
    try:
        for woche, klassen_plaene in plan_rohdaten.items():
            for klasse_name, tages_plaene in klassen_plaene.items():
                klasse_obj = klassen_map.get(klasse_name)
                if not klasse_obj: continue
                for tag, slots in tages_plaene.items():
                    for slot_idx, eintrag_data in enumerate(slots):
                        if eintrag_data:
                            angebot_obj = angebote_map.get(eintrag_data.get('angebot'))
                            lehrer1_obj = lehrer_map.get(eintrag_data.get('lehrer'))
                            lehrer2_obj = lehrer_map.get(eintrag_data.get('lehrer2')) if eintrag_data.get('lehrer2') else None
                            if not (angebot_obj and lehrer1_obj): continue
                            neuer_eintrag = StundenplanEintrag(
                                woche=woche, tag=tag, slot=slot_idx,
                                klasse_id=klasse_obj.id,
                                angebot_id=angebot_obj.id,
                                lehrer1_id=lehrer1_obj.id,
                                lehrer2_id=lehrer2_obj.id if lehrer2_obj else None
                            )
                            db.session.add(neuer_eintrag)
        db.session.commit()
        flash('Neuer Stundenplan erfolgreich erstellt und gespeichert!', 'success')
        session['scheduler_meldungen'] = meldungen
    except Exception as e:
        import traceback
        print("Fehler beim Speichern des Plans:", e)
        traceback.print_exc()
        db.session.rollback()
        flash(f'Fehler beim Speichern des neuen Plans: {e}', 'danger')
    return redirect(url_for('stundenplan_verwaltung'))
    # Benachrichtigung senden
    try:
        users_to_notify = User.query.filter(User.email.isnot(None)).all()
        if users_to_notify:
            recipients = [user.email for user in users_to_notify]
            html = render_template('email/stundenplan_notification.html')
            text = render_template('email/stundenplan_notification.txt')
            send_email("Stundenplan wurde aktualisiert", recipients, html, text)
            flash(f'Benachrichtigung über Plan-Update an {len(recipients)} Benutzer gesendet.', 'info')
    except Exception as e:
        flash(f"Fehler beim Senden der Benachrichtigungs-E-Mails: {e}", "warning")
    return redirect(url_for('stundenplan_verwaltung'))

def parse_zeiten(text):
    """Hilfsfunktion zum Parsen des Zeiten-Strings in eine strukturierte Liste."""
    zeiten = []
    for line in text.strip().split('\n'):
        line = line.strip()
        if "Pause" in line.lower():
            zeiten.append({"typ": "pause", "text": line})
        elif '-' in line:
            zeiten.append({"typ": "stunde", "text": line})
    return zeiten

# =====================================================================
# Benutzerverwaltung (Admin)
# =====================================================================

@app.route('/benutzer')
@login_required
@role_required([ROLE_ADMIN])
def benutzer_verwalten():
    """Rendert die Seite zur Verwaltung von Benutzern."""
    users = User.query.order_by(User.username).all()
    roles = [ROLE_ADMIN, ROLE_PLANER, ROLE_BENUTZER]
    # Farben für Auswahl im Formular
    user_farben = [
        "#3788d8", "#1f77b4", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2",
        "#7f7f7f", "#bcbd22", "#17becf", "#6c757d", "#ff7f0e", "#ffbb78", "#98df8a"
    ]
    return render_template('benutzer_verwalten.html', title='Benutzerverwaltung', users=users, roles=roles, user_farben=user_farben)


@app.route('/benutzer/save', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN])
def benutzer_save():
    """Speichert einen neuen oder aktualisiert einen bestehenden Benutzer."""
    user_id = request.form.get('user_id')
    username = request.form.get('username')
    email = request.form.get('email')
    role = request.form.get('role')
    password = request.form.get('password')
    farbe = request.form.get('farbe') or "#3788d8"

    if not username or not email or not role:
        flash('Benutzername, E-Mail und Rolle sind erforderlich.', 'danger')
        return redirect(url_for('benutzer_verwalten'))

    # Prüfen, ob Benutzername oder E-Mail bereits von einem ANDEREN Benutzer verwendet werden
    existing_user_by_name = User.query.filter(User.username == username, User.id != user_id).first()
    if existing_user_by_name:
        flash(f'Benutzername "{username}" ist bereits vergeben.', 'warning')
        return redirect(url_for('benutzer_verwalten'))

    existing_user_by_email = User.query.filter(User.email == email, User.id != user_id).first()
    if existing_user_by_email:
        flash(f'E-Mail "{email}" ist bereits vergeben.', 'warning')
        return redirect(url_for('benutzer_verwalten'))

    if user_id:  # Bearbeiten
        user = User.query.get_or_404(user_id)

        # Verhindern, dass der letzte Admin seine Rolle ändert
        if user.id == current_user.id and user.role == ROLE_ADMIN and role != ROLE_ADMIN:
            if User.query.filter_by(role=ROLE_ADMIN).count() <= 1:
                flash('Sie können Ihre Admin-Rolle nicht ändern, da Sie der einzige Admin sind.', 'danger')
                return redirect(url_for('benutzer_verwalten'))

        user.username = username
        user.email = email
        user.role = role
        user.farbe = farbe
        if password:
            user.password = bcrypt.generate_password_hash(password).decode('utf-8')
        flash(f'Benutzer "{username}" wurde aktualisiert.', 'success')
    else:
        if not password:
            flash('Ein Passwort ist für neue Benutzer erforderlich.', 'danger')
            return redirect(url_for('benutzer_verwalten'))
        
        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
        new_user = User(username=username, email=email, password=hashed_password, role=role, farbe=farbe)
        db.session.add(new_user)
        flash(f'Benutzer "{username}" wurde erstellt.', 'success')
    
    db.session.commit()
    return redirect(url_for('benutzer_verwalten'))

@app.route('/benutzer/delete/<int:user_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN])
def benutzer_delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('Sie können sich nicht selbst löschen.', 'danger')
        return redirect(url_for('benutzer_verwalten'))
    db.session.delete(user)
    db.session.commit()
    flash(f'Benutzer "{user.username}" wurde gelöscht.', 'success')
    return redirect(url_for('benutzer_verwalten'))

@app.route('/kalender')
def kalender():
    return render_template('kalender.html', title='Kalender')

def berechne_lehrer_uebersicht():
    """
    Berechnet eine detaillierte Übersicht der Soll- und Ist-Stunden für jeden Lehrer,
    aufgeschlüsselt nach A- und B-Woche und den jeweiligen Einsätzen.
    """
    lehrer_liste = Lehrer.query.order_by(Lehrer.name).all()
    plan_eintraege = StundenplanEintrag.query.options(
        db.joinedload(StundenplanEintrag.angebot),
        db.joinedload(StundenplanEintrag.klasse)
    ).all()

    # Datenstruktur für die Übersicht initialisieren
    lehrer_details = {
        lehrer.id: {
            "name": lehrer.name,
            "wochen": {
                "A": {"soll": lehrer.stunden_a or 0, "ist": 0, "einsaetze": defaultdict(int)},
                "B": {"soll": lehrer.stunden_b or 0, "ist": 0, "einsaetze": defaultdict(int)}
            }
        } for lehrer in lehrer_liste
    }

    # Ist-Stunden aus dem Stundenplan berechnen
    for eintrag in plan_eintraege:
        if not eintrag.angebot or not eintrag.klasse:
            continue

        einsatz_name = f"{eintrag.angebot.name} ({eintrag.klasse.name})"
        woche = eintrag.woche

        if eintrag.lehrer1_id and eintrag.lehrer1_id in lehrer_details:
            lehrer_details[eintrag.lehrer1_id]["wochen"][woche]["ist"] += 1
            lehrer_details[eintrag.lehrer1_id]["wochen"][woche]["einsaetze"][einsatz_name] += 1
        
        if eintrag.lehrer2_id and eintrag.lehrer2_id in lehrer_details:
            lehrer_details[eintrag.lehrer2_id]["wochen"][woche]["ist"] += 1
            lehrer_details[eintrag.lehrer2_id]["wochen"][woche]["einsaetze"][einsatz_name] += 1

    # Aufbereiten der Einsatzliste für das Template
    result = []
    for lehrer_id, details in lehrer_details.items():
        for woche in ["A", "B"]:
            einsaetze_liste = [{"name": name, "stunden": stunden} for name, stunden in details["wochen"][woche]["einsaetze"].items()]
            details["wochen"][woche]["einsaetze"] = sorted(einsaetze_liste, key=lambda x: x['name'])
        
        soll_gesamt = details["wochen"]["A"]["soll"] + details["wochen"]["B"]["soll"]
        ist_gesamt = details["wochen"]["A"]["ist"] + details["wochen"]["B"]["ist"]
        details["gesamt_diff"] = ist_gesamt - soll_gesamt
        details["id"] = lehrer_id
        result.append(details)

    return result

def berechne_klassen_uebersicht():
    """
    Berechnet eine detaillierte Übersicht der Soll- und Ist-Stunden für jede Klasse,
    aufgeschlüsselt nach Angeboten und Wochen.
    """
    klassen_liste = Klasse.query.order_by(Klasse.name).all()
    plan_eintraege = StundenplanEintrag.query.options(
        db.joinedload(StundenplanEintrag.angebot),
        db.joinedload(StundenplanEintrag.lehrer1),
        db.joinedload(StundenplanEintrag.lehrer2)
    ).all()

    # Datenstruktur für die Übersicht initialisieren
    klassen_details = {}
    for klasse in klassen_liste:
        klassen_details[klasse.id] = {
            "id": klasse.id,
            "name": klasse.name,
            "wochen": {
                "A": {"soll": 0, "ist": 0},
                "B": {"soll": 0, "ist": 0}
            },
            "angebote": defaultdict(lambda: {
                "name": "",
                "wochen": {
                    "A": {"soll": 0, "ist": 0},
                    "B": {"soll": 0, "ist": 0}
                },
                "lehrer": set()
            })
        }

        # Soll-Stunden aus Klassendefinition berechnen
        wochen_einstellung = klasse.woche or "AB"
        for angebot_def in klasse.angebote_stunden:
            angebot_name = angebot_def.get("angebot")
            stunden_gesamt = angebot_def.get("stunden_gesamt", 0)
            if not angebot_name:
                continue
            details_angebot = klassen_details[klasse.id]["angebote"][angebot_name]
            details_angebot["name"] = angebot_name
            if wochen_einstellung in ['A', 'AB']:
                details_angebot["wochen"]["A"]["soll"] += stunden_gesamt
                klassen_details[klasse.id]["wochen"]["A"]["soll"] += stunden_gesamt
            if wochen_einstellung in ['B', 'AB']:
                details_angebot["wochen"]["B"]["soll"] += stunden_gesamt
                klassen_details[klasse.id]["wochen"]["B"]["soll"] += stunden_gesamt

    # Ist-Stunden und Lehrer aus dem Stundenplan berechnen
    for eintrag in plan_eintraege:
        if eintrag.klasse_id not in klassen_details: continue
        klassen_details[eintrag.klasse_id]["wochen"][eintrag.woche]["ist"] += 1
        klassen_details[eintrag.klasse_id]["angebote"][eintrag.angebot.name]["wochen"][eintrag.woche]["ist"] += 1
        if eintrag.lehrer1: klassen_details[eintrag.klasse_id]["angebote"][eintrag.angebot.name]["lehrer"].add(eintrag.lehrer1.name)
        if eintrag.lehrer2: klassen_details[eintrag.klasse_id]["angebote"][eintrag.angebot.name]["lehrer"].add(eintrag.lehrer2.name)

    # Aufbereiten der Daten für das Template
    result = []
    for details in klassen_details.values():
        for woche in ["A", "B"]:
            details["wochen"][woche]["diff"] = details["wochen"][woche]["ist"] - details["wochen"][woche]["soll"]
        angebote_liste = []
        for angebot_name, angebot_details in sorted(details["angebote"].items()):
            for woche in ["A", "B"]:
                angebot_details["wochen"][woche]["diff"] = angebot_details["wochen"][woche]["ist"] - angebot_details["wochen"][woche]["soll"]
            angebot_details["lehrer"] = sorted(list(angebot_details["lehrer"]))
            angebote_liste.append(angebot_details)
        details["angebote"] = angebote_liste
        result.append(details)
    return result

@app.route('/uebersicht/lehrer')
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def lehrer_uebersicht():
    lehrer_details = berechne_lehrer_uebersicht()
    return render_template(
        'lehrer_uebersicht.html',
        title='Lehrer Übersicht',
        lehrer_details=lehrer_details,
    )

@app.route('/uebersicht/klassen')
@login_required
def klassen_uebersicht():
    klassen_details = berechne_klassen_uebersicht()
    return render_template(
        'klassen_uebersicht.html',
        title='Klassen Übersicht',
        klassen_details=klassen_details,
    )

# =====================================================================
# Einstellungen & Angebote
# =====================================================================

def get_setting(key, default=''):
    """Helper-Funktion, um eine Einstellung aus der DB zu lesen."""
    setting = Einstellung.query.filter_by(key=key).first()
    return setting.value if setting else default

def set_setting(key, value):
    """Helper-Funktion, um eine Einstellung in der DB zu speichern."""
    setting = Einstellung.query.filter_by(key=key).first()
    if setting:
        setting.value = value
    else:
        setting = Einstellung(key=key, value=value)
        db.session.add(setting)
    db.session.commit()

@app.route('/einstellungen')
@login_required
@role_required([ROLE_ADMIN]) # Einstellungen (Global) nur für Admin
def einstellungen():
    # Lade Einstellungen aus der DB
    zeiten_text = get_setting('zeiten_text', '08:00-08:45\n08:45-09:30\n09:30-09:45 Pause\n09:45-10:30\n10:30-11:15\n11:15-11:45 Pause\n11:45-12:30\n12:30-13:15')
    klassen_reihenfolge_a_str = get_setting('klassen_reihenfolge_a')
    klassen_reihenfolge_b_str = get_setting('klassen_reihenfolge_b')
    appearance_mode = get_setting('appearance_mode', 'System')
    stunden_pro_tag_str = get_setting('stunden_pro_tag', '{"Mo": 6, "Di": 6, "Mi": 6, "Do": 6, "Fr": 6}')
    stunden_pro_tag = json.loads(stunden_pro_tag_str)

    # Lade abhängige Daten
    alle_klassen = Klasse.query.order_by(Klasse.name).all()
    # Filtert Klassen, die in der jeweiligen Woche verfügbar sind (inkl. 'AB')
    klassen_fuer_woche_a = [k.name for k in alle_klassen if k.woche in ['A', 'AB']]
    klassen_fuer_woche_b = [k.name for k in alle_klassen if k.woche in ['B', 'AB']]
    
    # Intelligente Vorbelegung der Reihenfolge:
    # Nimm die gespeicherte Reihenfolge und füge neue, verfügbare Klassen am Ende hinzu.
    # Ist keine Reihenfolge gespeichert, nimm einfach alle verfügbaren Klassen.
    gespeicherte_klassen_a = [k for k in klassen_reihenfolge_a_str.split(',') if k]
    gespeicherte_klassen_b = [k for k in klassen_reihenfolge_b_str.split(',') if k]

    if not gespeicherte_klassen_a:
        final_order_a = klassen_fuer_woche_a
    else:
        final_order_a = gespeicherte_klassen_a + [k for k in klassen_fuer_woche_a if k not in gespeicherte_klassen_a]

    if not gespeicherte_klassen_b:
        final_order_b = klassen_fuer_woche_b
    else:
        final_order_b = gespeicherte_klassen_b + [k for k in klassen_fuer_woche_b if k not in gespeicherte_klassen_b]

    alle_angebote = Angebot.query.order_by(Angebot.name).all()

    return render_template('einstellungen.html',
                           title='Einstellungen',
                           zeiten_text=zeiten_text,
                           klassen_reihenfolge_a=final_order_a,
                           klassen_reihenfolge_b=final_order_b,
                           klassen_fuer_woche_a=klassen_fuer_woche_a,
                           klassen_fuer_woche_b=klassen_fuer_woche_b,
                           stunden_pro_tag=stunden_pro_tag,
                           tage_der_woche=TAGE_DER_WOCHE,
                           appearance_mode=appearance_mode,
                           alle_angebote=alle_angebote)

@app.route('/einstellungen_save', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN])
def einstellungen_save():
    set_setting('zeiten_text', request.form.get('zeiten_eingabe'))
    set_setting('klassen_reihenfolge_a', request.form.get('klassen_reihenfolge_a'))
    set_setting('klassen_reihenfolge_b', request.form.get('klassen_reihenfolge_b'))
    set_setting('appearance_mode', request.form.get('appearance_mode'))
    stunden_pro_tag_data = {
        tag: int(request.form.get(f'stunden_{tag.lower()}', 6)) for tag in TAGE_DER_WOCHE
    }
    set_setting('stunden_pro_tag', json.dumps(stunden_pro_tag_data))
    flash('Allgemeine Einstellungen gespeichert.', 'success')
    return redirect(url_for('einstellungen'))

@app.route('/angebot_save', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN])
def angebot_save():
    angebot_id = request.form.get('angebot_id')
    name = request.form.get('name')
    block_groesse = int(request.form.get('block_groesse', 2))
    nur_doppelblock = 'nur_ein_doppelblock' in request.form

    if not name:
        flash('Angebotsname darf nicht leer sein.', 'danger')
        return redirect(url_for('einstellungen'))

    if angebot_id: # Bearbeiten
        angebot = Angebot.query.get_or_404(angebot_id)
        angebot.name = name
        angebot.block_groesse = block_groesse
        angebot.nur_ein_doppelblock_pro_tag = nur_doppelblock
        flash(f'Angebot "{name}" aktualisiert.', 'success')
    else: # Neu
        existing = Angebot.query.filter_by(name=name).first()
        if existing:
            flash(f'Ein Angebot mit dem Namen "{name}" existiert bereits.', 'warning')
        else:
            neues_angebot = Angebot(name=name, block_groesse=block_groesse, nur_ein_doppelblock_pro_tag=nur_doppelblock)
            db.session.add(neues_angebot)
            flash(f'Angebot "{name}" hinzugefügt.', 'success')
    
    db.session.commit()
    return redirect(url_for('einstellungen'))

@app.route('/angebot_delete/<int:angebot_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN])
def angebot_delete(angebot_id):
    angebot = Angebot.query.get_or_404(angebot_id)
    # Sicherheitscheck: Ist das Angebot noch in Verwendung?
    if angebot.lehrer or StundenplanEintrag.query.filter_by(angebot_id=angebot.id).first():
        flash(f'Angebot "{angebot.name}" kann nicht gelöscht werden, da es noch von Lehrern oder im Stundenplan verwendet wird.', 'danger')
        return redirect(url_for('einstellungen'))
    
    db.session.delete(angebot)
    db.session.commit()
    flash(f'Angebot "{angebot.name}" wurde gelöscht.', 'success')
    return redirect(url_for('einstellungen'))

@app.route('/api/stundenplan/update', methods=['POST'])
def stundenplan_update():
    data = request.json
    try:
        print("DEBUG /api/stundenplan/update:", data)
        # Unterstütze beide Varianten: name/id für Klasse, Angebot, Lehrer
        woche = data.get("woche")
        klasse = (data.get("klasse") or data.get("klasse_name") or "").strip()
        tag = data.get("tag")
        slot = data.get("slot")
        angebot = data.get("angebot")
        angebot_id = data.get("angebot_id")
        lehrer1 = data.get("lehrer1")
        lehrer1_id = data.get("lehrer1_id")
        lehrer2 = data.get("lehrer2")
        lehrer2_id = data.get("lehrer2_id")
        force = data.get("force", False)

        # Robustere Prüfung: alle Felder müssen vorhanden und nicht leer sein
        missing = []
        if woche is None: missing.append("woche")
        if not klasse: missing.append("klasse/klasse_name")
        if tag is None: missing.append("tag")
        if slot is None: missing.append("slot")
        # Angebot und Lehrer1: entweder Name oder ID muss da sein
        if not angebot and not angebot_id: missing.append("angebot/angebot_id")
        if not lehrer1 and not lehrer1_id: missing.append("lehrer1/lehrer1_id")
        # lehrer2 kann leer sein
        if missing:
            return jsonify({"success": False, "error": f"Fehlende oder ungültige Felder: {', '.join(missing)}"}), 400

        # Hole Klasse
        klasse_obj = Klasse.query.filter_by(name=klasse).first()
        if not klasse_obj:
            return jsonify({"success": False, "error": "Klasse nicht gefunden"}), 400

        # Hole Angebot
        angebot_obj = None
        if angebot_id:
            angebot_obj = Angebot.query.filter_by(id=int(angebot_id)).first()
        elif angebot:
            angebot_obj = Angebot.query.filter_by(name=angebot).first()
        if not angebot_obj:
            return jsonify({"success": False, "error": "Angebot nicht gefunden"}), 400

        # Hole Lehrer1
        lehrer1_obj = None
        if lehrer1_id:
            lehrer1_obj = Lehrer.query.filter_by(id=int(lehrer1_id)).first()
        elif lehrer1:
            lehrer1_obj = Lehrer.query.filter_by(name=lehrer1).first()
        if not lehrer1_obj:
            return jsonify({"success": False, "error": "Lehrer1 nicht gefunden"}), 400

        # Hole Lehrer2 (optional)
        lehrer2_obj = None
        if lehrer2_id:
            lehrer2_obj = Lehrer.query.filter_by(id=int(lehrer2_id)).first()
        elif lehrer2:
            lehrer2_obj = Lehrer.query.filter_by(name=lehrer2).first()

        # --- PRÜFUNGEN AUF ÜBERSCHNEIDUNGEN UND ÜBERBUCHUNGEN ---
        warnungen = []
        konflikt_lehrer1 = StundenplanEintrag.query.filter(
            StundenplanEintrag.woche == woche,
            StundenplanEintrag.tag == tag,
            StundenplanEintrag.slot == int(slot),
            StundenplanEintrag.lehrer1_id == lehrer1_obj.id,
            StundenplanEintrag.klasse_id != klasse_obj.id
        ).first()
        if konflikt_lehrer1:
            warnungen.append(f"Lehrer {lehrer1_obj.name} ist zu dieser Zeit bereits in einer anderen Klasse eingetragen!")

        if lehrer2_obj:
            konflikt_lehrer2 = StundenplanEintrag.query.filter(
                StundenplanEintrag.woche == woche,
                StundenplanEintrag.tag == tag,
                StundenplanEintrag.slot == int(slot),
                StundenplanEintrag.lehrer1_id == lehrer2_obj.id,
                StundenplanEintrag.klasse_id != klasse_obj.id
            ).first()
            if konflikt_lehrer2:
                warnungen.append(f"Lehrer {lehrer2_obj.name} ist zu dieser Zeit bereits in einer anderen Klasse eingetragen!")

        max_stunden_tag_lehrer1 = lehrer1_obj.max_stunden_pro_tag.get(tag, {}).get(woche, 8)
        aktuelle_stunden_tag_lehrer1 = StundenplanEintrag.query.filter_by(
            woche=woche, tag=tag, lehrer1_id=lehrer1_obj.id
        ).count()
        if aktuelle_stunden_tag_lehrer1 >= max_stunden_tag_lehrer1:
            warnungen.append(f"Lehrer {lehrer1_obj.name} hat die maximale Stundenzahl am {tag} in Woche {woche} erreicht!")

        if lehrer2_obj:
            max_stunden_tag_lehrer2 = lehrer2_obj.max_stunden_pro_tag.get(tag, {}).get(woche, 8)
            aktuelle_stunden_tag_lehrer2 = StundenplanEintrag.query.filter_by(
                woche=woche, tag=tag, lehrer1_id=lehrer2_obj.id
            ).count()
            if aktuelle_stunden_tag_lehrer2 >= max_stunden_tag_lehrer2:
                warnungen.append(f"Lehrer {lehrer2_obj.name} hat die maximale Stundenzahl am {tag} in Woche {woche} erreicht!")

        klasse_max_stunden = klasse_obj.max_stunden_klasse or 6
        aktuelle_stunden_klasse = StundenplanEintrag.query.filter_by(
            woche=woche, tag=tag, klasse_id=klasse_obj.id
        ).count()
        if aktuelle_stunden_klasse >= klasse_max_stunden:
            warnungen.append(f"Klasse {klasse_obj.name} hat die maximale Stundenzahl am {tag} in Woche {woche} erreicht!")

        if warnungen and not force:
            # Hier wird success: false zurückgegeben, damit das Frontend das Popup zeigt
            return jsonify({"success": False, "error": "\n".join(warnungen)}), 200

        # --- Wenn force: true, dann wird gespeichert und success: true zurückgegeben ---
        # Slot und Tag ggf. konvertieren
        try:
            slot_int = int(slot) if not isinstance(slot, int) else slot
        except Exception:
            return jsonify({"success": False, "error": "Slot ungültig"}), 400

        # Suche bestehenden Eintrag
        eintrag = StundenplanEintrag.query.filter_by(
            woche=woche, tag=tag, slot=slot_int, klasse_id=klasse_obj.id
        ).first()
        if eintrag:
            eintrag.angebot_id = angebot_obj.id
            eintrag.lehrer1_id = lehrer1_obj.id
            eintrag.lehrer2_id = lehrer2_obj.id if lehrer2_obj else None
        else:
            eintrag = StundenplanEintrag(
                woche=woche,
                tag=tag,
                slot=slot_int,
                klasse_id=klasse_obj.id,
                angebot_id=angebot_obj.id,
                lehrer1_id=lehrer1_obj.id,
                lehrer2_id=lehrer2_obj.id if lehrer2_obj else None
            )
            db.session.add(eintrag)
        db.session.commit()
        # Seite soll nach dem Speichern neu geladen werden
        # Sende reload-Flag, damit das Frontend neu lädt
        return jsonify({"success": True, "reload": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/stundenplan/delete', methods=['POST'])
def stundenplan_delete():
    data = request.json
    try:
        print("DEBUG /api/stundenplan/delete:", data)

        # 1. Prüfe, ob das empfangene Objekt wirklich ein dict ist
        if not isinstance(data, dict):
            print("DELETE API: Kein dict empfangen!")
            return jsonify({"success": False, "error": "Ungültiges Datenformat"}), 400

        woche = data.get("woche")
        klasse = (data.get("klasse") or data.get("klasse_name") or "").strip()
        tag = data.get("tag")
        slot = data.get("slot")

        # 2. Zeige alle Keys und Werte im Request
        print("DELETE API: Keys im Request:", list(data.keys()))
        print(f"DELETE API: woche={woche!r} klasse={klasse!r} tag={tag!r} slot={slot!r}")
        print(f"Types: woche={type(woche)}, klasse={type(klasse)}, tag={type(tag)}, slot={type(slot)}")

        # 3. tag und slot können 0 sein, daher explizit auf None prüfen!
        missing = []
        if woche is None: missing.append("woche")
        if not klasse: missing.append("klasse/klasse_name")
        if tag is None: missing.append("tag")
        if slot is None: missing.append("slot")
        if missing:
            print("DELETE API: Fehlende Felder:", missing)
            return jsonify({"success": False, "error": f"Fehlende oder ungültige Felder: {', '.join(missing)}"}), 400

        # 4. Zeige alle Klassennamen aus der DB zum Vergleich
        alle_klassen = [k.name for k in Klasse.query.all()]
        print("DELETE API: Klassen in DB:", alle_klassen)
        klasse_obj = Klasse.query.filter_by(name=klasse).first()
        if not klasse_obj:
            print("DELETE API: Klasse nicht gefunden:", klasse)
            return jsonify({"success": False, "error": f"Klasse '{klasse}' nicht gefunden"}), 400

        # 5. tag und slot können Strings sein, aber auch schon int
        # --- Mapping für Tag-Strings zu Index ---
        tag_map = {"Mo": 0, "Di": 1, "Mi": 2, "Do": 3, "Fr": 4}
        try:
            # Falls tag ein Wochentag-String ist, wandle in Index um
            if isinstance(tag, str) and tag in tag_map:
                tag_int = tag_map[tag]
            else:
                tag_int = int(tag) if not isinstance(tag, int) else tag
            slot_int = int(slot) if not isinstance(slot, int) else slot
        except Exception:
            print("DELETE API: tag oder slot nicht konvertierbar:", tag, slot)
            return jsonify({"success": False, "error": "Tag oder Slot ungültig"}), 400

        print(f"DELETE API: Suche Eintrag mit woche={woche!r}, tag={tag_int!r}, slot={slot_int!r}, klasse_id={klasse_obj.id}")

        # --- Zusätzliche Debug-Ausgabe: Gibt es überhaupt einen Eintrag für diese Klasse? ---
        eintraege_klasse = StundenplanEintrag.query.filter_by(klasse_id=klasse_obj.id).all()
        print(f"DELETE API: Alle Einträge für Klasse {klasse} (id={klasse_obj.id}):")
        for e in eintraege_klasse:
            print(f"  woche={e.woche} tag={e.tag} slot={e.slot} angebot_id={e.angebot_id} lehrer1_id={e.lehrer1_id}")

        # --- Korrigierte Suche: tag als String vergleichen, nicht als Index ---
        gefundene = [
            e for e in eintraege_klasse
            if e.woche == woche and e.tag == tag and str(e.slot) == str(slot_int)
        ]
        print(f"DELETE API: Treffer für woche={woche}, tag={tag}, slot={slot_int}: {len(gefundene)}")
        for e in gefundene:
            print(f"  -> Eintrag: woche={e.woche} tag={e.tag} slot={e.slot} angebot_id={e.angebot_id} lehrer1_id={e.lehrer1_id}")

        # --- Suche nach Eintrag mit tag als String (z.B. "Mi") ---
        eintrag = StundenplanEintrag.query.filter_by(
            woche=woche, tag=tag, slot=slot_int, klasse_id=klasse_obj.id
        ).first()
        if eintrag:
            db.session.delete(eintrag)
            db.session.commit()
            print("DELETE API: Eintrag gelöscht.")
            # Seite soll nach dem Löschen neu geladen werden
            return jsonify({"success": True, "reload": True})
        else:
            print("DELETE API: Eintrag nicht gefunden für", woche, tag, slot_int, klasse_obj.id)
            return jsonify({"success": False, "error": "Eintrag nicht gefunden"}), 404
    except Exception as e:
        db.session.rollback()
        print("DELETE API: Exception:", str(e))
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/events', methods=['GET'])
@login_required
def api_events():
    events = []
    for event in Event.query.options(db.joinedload(Event.user)).all():
        # --- NEU: Logik für status-basierte Anzeige ---
        event_title = event.title
        background_color = event.user.farbe if hasattr(event.user, "farbe") else "#3788d8"
        
        if event.status == 'pending':
            event_title = f"[Offen] {event.title}"
            background_color = "#f39c12" # Gelb
        elif event.status == 'rejected':
            event_title = f"[Abgelehnt] {event.title}"
            background_color = "#e74c3c" # Rot
        
        # Admins/Planer sehen den Ersteller-Namen
        if current_user.role in [ROLE_ADMIN, ROLE_PLANER]:
             event_title += f" ({event.user.username})"

        events.append({
            "id": event.id,
            "title": event_title,
            "raw_title": event.title,
            "start": event.start.isoformat(),
            "end": event.end.isoformat() if event.end else None,
            "allDay": event.all_day,
            "backgroundColor": background_color,
            "borderColor": background_color,
            "status": event.status,
            "creator_id": event.user_id,
            "creator_name": event.user.username
        })
    return jsonify(events)

@app.route('/api/events/add', methods=['POST'])
@login_required
def api_events_add():
    data = request.get_json()
    title = data.get('title')
    start = data.get('start')
    end = data.get('end')
    all_day = data.get('allDay', False)
    if not title or not start:
        return jsonify({"status": "error", "message": "Titel und Startzeit sind erforderlich."}), 400
    event = Event(
        title=title,
        start=datetime.fromisoformat(start.replace('Z', '+00:00')),
        end=datetime.fromisoformat(end.replace('Z', '+00:00')) if end else None,
        all_day=all_day,
        user_id=current_user.id
    )
    db.session.add(event)
    db.session.commit()

    # Benachrichtigung an Admins und Planer senden
    try:
        users_to_notify = User.query.filter(User.role.in_([ROLE_ADMIN, ROLE_PLANER]), User.email.isnot(None)).all()
        if users_to_notify:
            recipients = [user.email for user in users_to_notify]
            html = render_template('email/kalender_notification.html', event=event, creator=current_user)
            text = render_template('email/kalender_notification.txt', event=event, creator=current_user)
            send_email(f"Neuer Kalendereintrag: {event.title}", recipients, html, text)
    except Exception as e:       
        print(f"Fehler beim Senden der Kalender-Benachrichtigung: {e}")
    return jsonify({"status": "success", "id": event.id, "message": "Event erstellt"})
    #email Benachrichtigungen
    try:
        users_to_notify = User.query.filter(User.role.in_([ROLE_ADMIN, ROLE_PLANER]), User.email.isnot(None)).all()
        recipients = [user.email for user in users_to_notify]
        html = render_template('email/kalender_notification.html', event=event, creator=current_user)
        text = render_template('email/kalender_notification.txt', event=event, creator=current_user)
        send_email(f"Neuer Kalendereintrag: {event.title}", recipients, html, text)
    except Exception as e:
        print(f"Fehler beim Senden der Kalender-Benachrichtigung: {e}")

@app.route('/api/events/update/<int:event_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def api_events_update(event_id):
    event = Event.query.get_or_404(event_id)
    data = request.get_json()
    event.title = data.get('title', event.title)
    event.start = datetime.fromisoformat(data.get('start').replace('Z', '+00:00'))
    event.end = datetime.fromisoformat(data.get('end').replace('Z', '+00:00')) if data.get('end') else None
    event.all_day = data.get('allDay', False)
    db.session.commit()
    return jsonify({"status": "success"})

@app.route('/api/events/delete/<int:event_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def api_events_delete(event_id):
    event = Event.query.get_or_404(event_id)
    db.session.delete(event)
    db.session.commit()
    return jsonify({"status": "success"})

@app.route('/api/events/approve/<int:event_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def api_event_approve(event_id):
    event = Event.query.get_or_404(event_id)
    event.status = 'approved'
    db.session.commit()
    # --- NEU: Benachrichtigung senden ---
    try:
        if event.user and event.user.email:
            html = render_template('email/event_approved.html', event=event, creator=event.user)
            text = render_template('email/event_approved.txt', event=event, creator=event.user)
            send_email(f"Dein Termin wurde genehmigt: {event.title}", [event.user.email], html, text)
    except Exception as e:
        app.logger.error(f"Fehler beim Senden der Genehmigungs-E-Mail: {e}")
    return jsonify({"status": "success", "message": "Termin genehmigt."})

@app.route('/api/events/reject/<int:event_id>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER])
def api_event_reject(event_id):
    event = Event.query.get_or_404(event_id)
    event.status = 'rejected'
    db.session.commit()
    # --- NEU: Benachrichtigung senden ---
    try:
        if event.user and event.user.email:
            html = render_template('email/event_rejected.html', event=event, creator=event.user)
            text = render_template('email/event_rejected.txt', event=event, creator=event.user)
            send_email(f"Dein Termin wurde abgelehnt: {event.title}", [event.user.email], html, text)
    except Exception as e:
        app.logger.error(f"Fehler beim Senden der Ablehnungs-E-Mail: {e}")
    return jsonify({"status": "success", "message": "Termin abgelehnt."})

def style_worksheet(ws, header_fill_color="4F81BD"):
    """
    Formatiert ein komplettes Tabellenblatt mit professionellen Stilen.
    - Formatiert die Kopfzeile (fett, farbiger Hintergrund).
    - Fügt allen Zellen einen Rahmen hinzu.
    - Passt die Spaltenbreite automatisch an den Inhalt an.
    """
    # Stile definieren
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Kopfzeile (erste Zeile) formatieren
    if ws.max_row > 0:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = header_alignment

    # Datenzellen formatieren
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = cell_alignment

    # Spaltenbreiten anpassen
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    cell_max_line_length = max(len(line) for line in str(cell.value).split('\n'))
                    if cell_max_line_length > max_length:
                        max_length = cell_max_line_length
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

@app.route('/excel_export', methods=['POST'])
def excel_export():
    # Angebote
    angebote = Angebot.query.order_by(Angebot.name).all()
    angebote_dicts = [
        {
            "ID": a.id,
            "Name": a.name,
            "Blockgröße": a.block_groesse,
            "Nur ein Doppelblock pro Tag": "Ja" if a.nur_ein_doppelblock_pro_tag else "Nein"
        }
        for a in angebote
    ]

    # Lehrer Übersicht
    lehrer_uebersicht = berechne_lehrer_uebersicht()
    lehrer_rows = []
    for l in lehrer_uebersicht:
        for woche in ["A", "B"]:
            for einsatz in l["wochen"][woche]["einsaetze"]:
                lehrer_rows.append({
                    "Lehrer": l["name"],
                    "Woche": woche,
                    "Soll": l["wochen"][woche]["soll"],
                    "Ist": l["wochen"][woche]["ist"],
                    "Einsatz": einsatz["name"],
                    "Einsatz-Stunden": einsatz["stunden"],
                    "Gesamt-Differenz": l["gesamt_diff"]
                })
            # Falls keine Einsätze, trotzdem Zeile
            if not l["wochen"][woche]["einsaetze"]:
                lehrer_rows.append({
                    "Lehrer": l["name"],
                    "Woche": woche,
                    "Soll": l["wochen"][woche]["soll"],
                    "Ist": l["wochen"][woche]["ist"],
                    "Einsatz": "",
                    "Einsatz-Stunden": "",
                    "Gesamt-Differenz": l["gesamt_diff"]
                })

    # Klassen Übersicht
    klassen_uebersicht = berechne_klassen_uebersicht()
    klassen_rows = []
    for k in klassen_uebersicht:
        for angebot in k["angebote"]:
            for woche in ["A", "B"]:
                klassen_rows.append({
                    "Klasse": k["name"],
                    "Angebot": angebot["name"],
                    "Woche": woche,
                    "Soll": angebot["wochen"][woche]["soll"],
                    "Ist": angebot["wochen"][woche]["ist"],
                    "Differenz": angebot["wochen"][woche]["diff"],
                    "Lehrer": ", ".join(angebot["lehrer"])
                })

    # Stundenplan (alle Einträge)
    stundenplan_eintraege = StundenplanEintrag.query.all()
    stundenplan_rows = []
    for e in stundenplan_eintraege:
        stundenplan_rows.append({
            "Woche": e.woche,
            "Tag": e.tag,
            "Slot": e.slot,
            "Klasse": e.klasse.name if e.klasse else "",
            "Angebot": e.angebot.name if e.angebot else "",
            "Lehrer 1": e.lehrer1.name if e.lehrer1 else "",
            "Lehrer 2": e.lehrer2.name if e.lehrer2 else ""
        })

    # Excel-Datei im Speicher erstellen und formatieren
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Standard-Blatt entfernen

    # Tabellenblatt für Angebote
    ws_angebote = wb.create_sheet("Angebote")
    if angebote_dicts:
        ws_angebote.append(list(angebote_dicts[0].keys()))
        for row in angebote_dicts: ws_angebote.append(list(row.values()))
    style_worksheet(ws_angebote, header_fill_color="767171")

    # Tabellenblatt für Lehrer

    ws_lehrer = wb.create_sheet("Lehrer Übersicht")
    if lehrer_rows:
        ws_lehrer.append(list(lehrer_rows[0].keys()))
        for row in lehrer_rows: ws_lehrer.append(list(row.values()))
    style_worksheet(ws_lehrer, header_fill_color="4F81BD")

    # Tabellenblatt für Klassen
    ws_klassen = wb.create_sheet("Klassen Übersicht")
    if klassen_rows:
        ws_klassen.append(list(klassen_rows[0].keys()))
        for row in klassen_rows: ws_klassen.append(list(row.values()))
    style_worksheet(ws_klassen, header_fill_color="375623")

    # Tabellenblatt für Stundenplan
    ws_stundenplan = wb.create_sheet("Stundenplan")
    if stundenplan_rows:
        ws_stundenplan.append(list(stundenplan_rows[0].keys()))
        for row in stundenplan_rows: ws_stundenplan.append(list(row.values()))
    style_worksheet(ws_stundenplan, header_fill_color="C0504D")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="Stundenplan_Export.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Nein, an app.py muss nichts angepasst werden,
# solange du die Rückgabe von scheduler_web.generate_schedule_data wie bisher verwendest
# und die Meldungen (session['scheduler_meldungen'] oder flash) im Frontend anzeigst.

# Die Logik für das Filtern und Überspringen unplanbarer Anforderungen ist jetzt komplett in scheduler_web.py.
# Die Schnittstelle und das Rückgabeformat bleiben gleich.

# Optional: du kannst im Frontend die self.meldungen als Popup oder als Liste anzeigen,
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.
# Optional: du kannst im Frontend die self.meldungen als Popup oder als Liste anzeigen,
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.
# Optional: du kannst im Frontend die self.meldungen als Popup oder als Liste anzeigen,
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.
# Optional: du kannst im Frontend die self.meldungen als Popup oder als Liste anzeigen,
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.
# Die Schnittstelle und das Rückgabeformat bleiben gleich.

# Optional: du kannst im Frontend die self.meldungen als Popup oder als Liste anzeigen,
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.
# Optional: du kannst im Frontend die self.meldungen als Popup oder als Liste anzeigen,
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.
# Optional: du kannst im Frontend die self.meldungen als Popup oder als Liste anzeigen,
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.
# Optional: du kannst im Frontend die self.meldungen als Popup oder als Liste anzeigen,
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.
# damit der Nutzer sofort sieht, welche Anforderungen nicht planbar waren.

@app.route('/profil', methods=['GET', 'POST'])
@login_required
def profil():
    user_farben = [
        "#3788d8", "#1f77b4", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2",
        "#7f7f7f", "#bcbd22", "#17becf", "#6c757d", "#ff7f0e", "#ffbb78", "#98df8a"
    ]
    if request.method == 'POST':
        farbe = request.form.get('farbe') or "#3788d8"
        password = request.form.get('password')
        if farbe:
            current_user.farbe = farbe
        if password:
            current_user.password = bcrypt.generate_password_hash(password).decode('utf-8')
        db.session.commit()
        flash("Profil gespeichert.", "success")
        return redirect(url_for('profil'))
    return render_template('profil.html', user_farben=user_farben)

# =====================================================================
# Datei-Downloads
# =====================================================================

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'downloads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'xlsx', 'xls', 'csv', 'txt', 'zip', 'png', 'jpg', 'jpeg'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/downloads')
@login_required
def downloads():
    files = sorted(os.listdir(UPLOAD_FOLDER))
    return render_template('downloads.html', files=files)

@app.route('/downloads/upload', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN])
def downloads_upload():
    if 'file' not in request.files:
        flash('Keine Datei ausgewählt.', 'danger')
        return redirect(url_for('downloads'))
    file = request.files['file']
    if file.filename == '':
        flash('Keine Datei ausgewählt.', 'danger')
        return redirect(url_for('downloads'))
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        flash('Datei erfolgreich hochgeladen.', 'success')
    else:
        flash('Dateityp nicht erlaubt.', 'danger')
    return redirect(url_for('downloads'))

@app.route('/downloads/delete/<filename>', methods=['POST'])
@login_required
@role_required([ROLE_ADMIN])
def downloads_delete(filename):
    path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(path):
        os.remove(path)
        flash('Datei gelöscht.', 'success')
    else:
        flash('Datei nicht gefunden.', 'danger')
    return redirect(url_for('downloads'))

@app.route('/downloads/files/<filename>')
@login_required
def downloads_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

# =====================================================================
# Kanban Board
# =====================================================================

@app.route('/kanban')
@login_required
def kanban_board():
    """Displays the Kanban board."""
    lists = KanbanList.query.order_by(KanbanList.position).all()
    return render_template('kanban.html', title='Aufgaben (Kanban)', lists=lists)

@app.route('/api/kanban/card/add', methods=['POST'])
@login_required
def kanban_add_card():
    data = request.get_json()
    content = data.get('content')
    list_id = data.get('list_id')

    if not content or not list_id:
        return jsonify({'success': False, 'error': 'Fehlender Inhalt oder Listen-ID'}), 400

    kanban_list = KanbanList.query.get(list_id)
    if not kanban_list:
        return jsonify({'success': False, 'error': 'Liste nicht gefunden'}), 404

    # Karte am Ende der Liste hinzufügen
    position = len(kanban_list.cards)

    new_card = KanbanCard(
        content=content,
        list_id=list_id,
        position=position,
        user_id=current_user.id
    )
    db.session.add(new_card)
    db.session.commit()

    return jsonify({'success': True, 'card_id': new_card.id})

@app.route('/api/kanban/card/move', methods=['POST'])
@login_required
def kanban_move_card():
    data = request.get_json()
    new_list_id = data.get('new_list_id')
    ordered_ids = data.get('ordered_ids') # Array mit Karten-IDs in der neuen Reihenfolge

    if not new_list_id or ordered_ids is None:
        return jsonify({'success': False, 'error': 'Fehlende Parameter'}), 400

    # Aktualisiere alle Karten in der Zielliste, um die neue Reihenfolge widerzuspiegeln
    for index, card_id in enumerate(ordered_ids):
        card = KanbanCard.query.get(card_id)
        if card:
            card.list_id = new_list_id
            card.position = index
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/kanban/card/delete/<int:card_id>', methods=['POST'])
@login_required
def kanban_delete_card(card_id):
    card = KanbanCard.query.get_or_404(card_id)

    # Optional: Hier könnte eine Berechtigungsprüfung stehen
    # z.B. ob der current_user der Ersteller ist oder Admin-Rechte hat.
    # if card.user_id != current_user.id and current_user.role != ROLE_ADMIN:
    #     return jsonify({'success': False, 'error': 'Keine Berechtigung'}), 403

    db.session.delete(card)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/kanban/card/edit/<int:card_id>', methods=['POST'])
@login_required
def kanban_edit_card(card_id):
    card = KanbanCard.query.get_or_404(card_id)
    data = request.get_json()
    new_content = data.get('content')

    # Berechtigungsprüfung: Nur Ersteller oder Admin dürfen bearbeiten
    if card.user_id != current_user.id and current_user.role != ROLE_ADMIN:
        return jsonify({'success': False, 'error': 'Keine Berechtigung zum Bearbeiten'}), 403

    if not new_content or not new_content.strip():
        return jsonify({'success': False, 'error': 'Inhalt darf nicht leer sein'}), 400

    card.content = new_content.strip()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Karte aktualisiert'})


@app.route('/verwaltung/vertretungsplan')
@login_required
# @admin_oder_planer_required # Fügen Sie Ihren Berechtigungs-Decorator hinzu
def vertretungsplan_verwalten():
    """Zeigt die Verwaltungsseite für den Vertretungsplan an."""
    plan = Vertretungsplan.query.first()

    return render_template('vertretungsplan_verwalten.html', 
                           title="Vertretungsplan verwalten",
                           aktiver_plan=plan)


@app.route('/vertretungsplan/erstellen', methods=['POST'])
@login_required
# @admin_oder_planer_required # Fügen Sie Ihren Berechtigungs-Decorator hinzu
def vertretungsplan_erstellen():
    """Erstellt einen neuen Vertretungsplan als Kopie des Haupt-Stundenplans."""
    gueltig_von_str = request.form.get('gueltig_von')
    gueltig_bis_str = request.form.get('gueltig_bis')
    vorlage_woche = request.form.get('vorlage_woche')

    try:
        gueltig_von = datetime.strptime(gueltig_von_str, '%Y-%m-%d').date()
        gueltig_bis = datetime.strptime(gueltig_bis_str, '%Y-%m-%d').date()

        if gueltig_von > gueltig_bis:
            flash('Das "Gültig von"-Datum darf nicht nach dem "Gültig bis"-Datum liegen.', 'danger')
            return redirect(url_for('vertretungsplan_verwalten'))

        # 1. Bestehenden Plan löschen
        Vertretungsplan.query.delete()
        
        # 2. Neuen Plan-Container erstellen
        neuer_plan = Vertretungsplan(gueltig_von=gueltig_von, gueltig_bis=gueltig_bis, vorlage_woche=vorlage_woche)
        db.session.add(neuer_plan)
        db.session.flush()  # Nötig, um die ID für die Einträge zu bekommen

        # 3. Einträge aus der Vorlage kopieren
        vorlage_eintraege = StundenplanEintrag.query.filter_by(woche=vorlage_woche).all()
        if not vorlage_eintraege:
            db.session.rollback()
            flash(f'Vorlage (Woche {vorlage_woche}) ist leer. Kein Vertretungsplan erstellt.', 'warning')
            return redirect(url_for('vertretungsplan_verwalten'))

        for eintrag in vorlage_eintraege:
            neuer_eintrag = VertretungsplanEintrag(
                vertretungsplan_id=neuer_plan.id,
                tag=eintrag.tag,
                slot=eintrag.slot,
                klasse_id=eintrag.klasse_id,
                angebot_id=eintrag.angebot_id,
                lehrer1_id=eintrag.lehrer1_id,
                lehrer2_id=eintrag.lehrer2_id
            )
            db.session.add(neuer_eintrag)

        db.session.commit()
        flash('Vertretungsplan wurde erfolgreich erstellt.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Ein Fehler ist aufgetreten: {str(e)}', 'danger')
        app.logger.error(f"Fehler beim Erstellen des Vertretungsplans: {e}")

    return redirect(url_for('vertretungsplan_verwalten'))


@app.route('/vertretungsplan/loeschen', methods=['POST'])
@login_required
# @admin_oder_planer_required # Fügen Sie Ihren Berechtigungs-Decorator hinzu
def vertretungsplan_loeschen():
    """Löscht den aktuell aktiven Vertretungsplan."""
    try:
        Vertretungsplan.query.delete()
        db.session.commit()
        flash('Der aktive Vertretungsplan wurde erfolgreich gelöscht.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Löschen des Vertretungsplans: {str(e)}', 'danger')
        app.logger.error(f"Fehler beim Löschen des Vertretungsplans: {e}")
        
    return redirect(url_for('vertretungsplan_verwalten'))


@app.route('/vertretungsplan/bearbeiten/<int:plan_id>')
@login_required
# @admin_oder_planer_required # Berechtigungsprüfung nicht vergessen
def vertretungsplan_bearbeiten(plan_id):
    """Zeigt die Seite zum Bearbeiten eines Vertretungsplans an."""
    plan = Vertretungsplan.query.get_or_404(plan_id)
    
    # Lade alle Daten, die für die Anzeige des Plans benötigt werden.
    # Diese Logik ist sehr ähnlich zur stundenplan_verwaltung.
    
    # --- NEU: Klassen basierend auf der Vorlagen-Woche des Plans filtern ---
    woche = plan.vorlage_woche
    klassen_reihenfolge_a_str = get_setting('klassen_reihenfolge_a', '')
    klassen_reihenfolge_b_str = get_setting('klassen_reihenfolge_b', '')
    alle_klassen_map = {k.name: k for k in Klasse.query.all()}
    
    if woche == 'A':
        geordnete_klassen_namen = klassen_reihenfolge_a_str.split(',')
    else:
        geordnete_klassen_namen = klassen_reihenfolge_b_str.split(',')
        
    geordnete_klassen = [alle_klassen_map[name] for name in geordnete_klassen_namen if name in alle_klassen_map]
    
    zeiten_text = get_setting('zeiten_text', '08:00-08:45\n08:45-09:30\n09:30-09:45 Pause\n09:45-10:30\n10:30-11:15\n11:15-11:45 Pause\n11:45-12:30\n12:30-13:15')
    zeit_slots = parse_zeiten(zeiten_text)
    num_slots = len(zeit_slots)

    plan_eintraege = VertretungsplanEintrag.query.filter_by(vertretungsplan_id=plan.id).all()
    plan_data_per_tag = defaultdict(lambda: defaultdict(dict))
    for eintrag in plan_eintraege:
        if not all([eintrag.angebot, eintrag.lehrer1, eintrag.klasse]):
            continue
        eintrag_dict = {
            'angebot': {'id': eintrag.angebot.id, 'name': eintrag.angebot.name},
            'lehrer1': {'id': eintrag.lehrer1.id, 'name': eintrag.lehrer1.name, 'farbe': eintrag.lehrer1.farbe},
            'lehrer2': {'id': eintrag.lehrer2.id, 'name': eintrag.lehrer2.name, 'farbe': eintrag.lehrer2.farbe} if eintrag.lehrer2 else None
        }
        if eintrag.slot < num_slots:
            plan_data_per_tag[eintrag.tag][eintrag.slot][eintrag.klasse.name] = eintrag_dict

    return render_template('vertretungsplan_bearbeiten.html',
                           title=f"Vertretungsplan bearbeiten",
                           plan=plan,
                           geordnete_klassen=geordnete_klassen,
                           plan_data_per_tag=plan_data_per_tag,
                           zeit_slots=zeit_slots,
                           tage_der_woche=TAGE_DER_WOCHE,
                           alle_lehrer=Lehrer.query.order_by(Lehrer.name).all(),
                           alle_angebote=Angebot.query.order_by(Angebot.name).all()
                           )

@app.route('/api/vertretungsplan/update', methods=['POST'])
@login_required
# @admin_oder_planer_required
def vertretungsplan_update():
    """API-Endpunkt zum Aktualisieren eines Eintrags im Vertretungsplan."""
    data = request.json
    try:
        plan_id = data.get("plan_id")
        klasse_name = data.get("klasse_name")
        tag = data.get("tag")
        slot = int(data.get("slot"))
        angebot_id = int(data.get("angebot_id"))
        lehrer1_id = int(data.get("lehrer1_id"))
        lehrer2_id = int(data.get("lehrer2_id")) if data.get("lehrer2_id") else None

        klasse_obj = Klasse.query.filter_by(name=klasse_name).first()
        if not klasse_obj:
            return jsonify({"success": False, "error": "Klasse nicht gefunden"}), 404

        # Finde bestehenden Eintrag oder erstelle einen neuen
        eintrag = VertretungsplanEintrag.query.filter_by(
            vertretungsplan_id=plan_id, tag=tag, slot=slot, klasse_id=klasse_obj.id
        ).first()

        if eintrag:
            eintrag.angebot_id = angebot_id
            eintrag.lehrer1_id = lehrer1_id
            eintrag.lehrer2_id = lehrer2_id
        else:
            eintrag = VertretungsplanEintrag(
                vertretungsplan_id=plan_id,
                tag=tag,
                slot=slot,
                klasse_id=klasse_obj.id,
                angebot_id=angebot_id,
                lehrer1_id=lehrer1_id,
                lehrer2_id=lehrer2_id
            )
            db.session.add(eintrag)
        
        db.session.commit()
        return jsonify({"success": True, "reload": True})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Fehler beim Update des Vertretungsplans: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/vertretungsplan/delete', methods=['POST'])
@login_required
# @admin_oder_planer_required
def vertretungsplan_delete_entry():
    """API-Endpunkt zum Löschen eines Eintrags aus dem Vertretungsplan."""
    data = request.json
    try:
        plan_id = data.get("plan_id")
        klasse_name = data.get("klasse_name")
        tag = data.get("tag")
        slot = int(data.get("slot"))

        klasse_obj = Klasse.query.filter_by(name=klasse_name).first()
        if klasse_obj:
            VertretungsplanEintrag.query.filter_by(
                vertretungsplan_id=plan_id, tag=tag, slot=slot, klasse_id=klasse_obj.id
            ).delete()
            db.session.commit()
        
        return jsonify({"success": True, "reload": True})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Fehler beim Löschen eines Vertretungsplan-Eintrags: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


# --- NEUE ROUTE für die schreibgeschützte Ansicht ---
@app.route('/vertretungsplan/ansicht/<int:plan_id>')
@login_required
@role_required([ROLE_ADMIN, ROLE_PLANER, ROLE_BENUTZER]) # Alle dürfen sehen
def vertretungsplan_anzeigen_readonly(plan_id):
    """Zeigt eine schreibgeschützte Ansicht eines Vertretungsplans."""
    plan = Vertretungsplan.query.get_or_404(plan_id)
    
    # Logik zum Laden der Daten ist identisch zur Bearbeitungs-Route
    woche = plan.vorlage_woche
    klassen_reihenfolge_a_str = get_setting('klassen_reihenfolge_a', '')
    klassen_reihenfolge_b_str = get_setting('klassen_reihenfolge_b', '')
    alle_klassen_map = {k.name: k for k in Klasse.query.all()}
    
    if woche == 'A':
        geordnete_klassen_namen = klassen_reihenfolge_a_str.split(',')
    else:
        geordnete_klassen_namen = klassen_reihenfolge_b_str.split(',')
        
    geordnete_klassen = [alle_klassen_map[name] for name in geordnete_klassen_namen if name in alle_klassen_map]
    
    zeiten_text = get_setting('zeiten_text', '08:00-08:45\n08:45-09:30\n09:30-09:45 Pause\n09:45-10:30\n10:30-11:15\n11:15-11:45 Pause\n11:45-12:30\n12:30-13:15')
    zeit_slots = parse_zeiten(zeiten_text)
    num_slots = len(zeit_slots)

    plan_eintraege = VertretungsplanEintrag.query.filter_by(vertretungsplan_id=plan.id).all()
    plan_data_per_tag = defaultdict(lambda: defaultdict(dict))
    for eintrag in plan_eintraege:
        if not all([eintrag.angebot, eintrag.lehrer1, eintrag.klasse]):
            continue
        eintrag_dict = {
            'angebot': {'id': eintrag.angebot.id, 'name': eintrag.angebot.name},
            'lehrer1': {'id': eintrag.lehrer1.id, 'name': eintrag.lehrer1.name, 'farbe': eintrag.lehrer1.farbe},
            'lehrer2': {'id': eintrag.lehrer2.id, 'name': eintrag.lehrer2.name, 'farbe': eintrag.lehrer2.farbe} if eintrag.lehrer2 else None
        }
        if eintrag.slot < num_slots:
            plan_data_per_tag[eintrag.tag][eintrag.slot][eintrag.klasse.name] = eintrag_dict

    return render_template('vertretungsplan_ansicht.html',
                           title=f"Vertretungsplan (Ansicht)",
                           plan=plan,
                           geordnete_klassen=geordnete_klassen,
                           plan_data_per_tag=plan_data_per_tag,
                           zeit_slots=zeit_slots,
                           tage_der_woche=TAGE_DER_WOCHE,
                           alle_lehrer=Lehrer.query.order_by(Lehrer.name).all()
                           )